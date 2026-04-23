"""날짜별 대시보드 — 자재 상·하차 반입/반출 현황 테이블."""
import io
import sqlite3
from datetime import date, timedelta
import streamlit as st
from config import KIND_IN, KIND_OUT
from db.models import settings_get


_DASH_CSS = """
<style>
/* 날짜 네비 */
.st-key-dash_nav { margin-bottom: 16px !important; }
.st-key-dash_nav .stHorizontalBlock {
  gap: 4px !important;
  align-items: center !important;
  flex-wrap: nowrap !important;
}
/* 버튼 컬럼: 고정 최소폭 */
.st-key-dash_nav .stHorizontalBlock > div:nth-child(1),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(2),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(4),
.st-key-dash_nav .stHorizontalBlock > div:nth-child(5) {
  flex: 0 0 40px !important;
  min-width: 40px !important;
  max-width: 40px !important;
}
/* 날짜 박스 컬럼: 남은 공간 차지 */
.st-key-dash_nav .stHorizontalBlock > div:nth-child(3) {
  flex: 1 1 auto !important;
  min-width: 0 !important;
}
.st-key-dash_nav button {
  height: 38px !important;
  min-height: 38px !important;
  max-height: 38px !important;
  padding: 0 !important;
  font-size: 13px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
}
.st-key-dash_nav button p,
.st-key-dash_nav button span {
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  margin: 0 !important;
  line-height: 1 !important;
}
.st-key-dash_nav [data-baseweb="input"] {
  height: 38px !important;
  min-height: 38px !important;
}
.st-key-dash_nav [data-baseweb="input"] input {
  height: 38px !important;
  line-height: 38px !important;
  padding-top: 0 !important;
  padding-bottom: 0 !important;
  text-align: center !important;
}

/* 대시보드 래퍼 */
.dash-wrap {
  width: 100%;
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}

/* 타이틀 박스 */
.dash-title-box {
  border: 2px solid #1e3a8a;
  border-radius: 4px;
  text-align: center;
  padding: 16px 8px 14px 8px;
  margin-bottom: 12px;
  width: 100%;
  box-sizing: border-box;
}
.dash-title-box h2 {
  font-size: clamp(20px, 4vw, 32px);
  font-weight: 900;
  color: #0f172a;
  margin: 0;
  letter-spacing: -0.5px;
}

/* 사이트·날짜 헤더 */
.dash-meta {
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 10px;
  font-size: 13px;
  color: #1e3a8a;
  font-weight: 600;
}

/* 테이블 */
.dash-table {
  width: 100%;
  border-collapse: collapse;
  font-size: clamp(11px, 1.8vw, 13px);
  min-width: 860px;
}
.dash-table th {
  background: #1e3a8a;
  color: #ffffff;
  padding: 8px 6px;
  text-align: center;
  border: 1px solid #1e40af;
  font-weight: 700;
  white-space: nowrap;
  vertical-align: middle;
}
.dash-table th.th-sub {
  background: #2563eb;
  font-size: 11px;
  padding: 5px 4px;
}
.dash-table td {
  padding: 7px 6px;
  border: 1px solid #cbd5e1;
  text-align: center !important;
  vertical-align: middle !important;
  color: #0f172a;
  word-break: keep-all;
}
.dash-table tr:nth-child(even) td { background: #f8fafc; }
.dash-table tr:nth-child(odd)  td { background: #ffffff; }
.dash-table tr:hover td { background: #eff6ff !important; }

/* 반입/반출 배지 */
.kind-in  { color: #1d4ed8; font-weight: 700; }
.kind-out { color: #b91c1c; font-weight: 700; }

/* 합계 행 */
.dash-table tr.total-row td {
  background: #e0e7ff !important;
  font-weight: 700;
  color: #1e3a8a;
}

/* 다운로드 버튼 */
[data-testid="stDownloadButton"] button {
  white-space: nowrap !important;
  background-color: #2563eb !important;
  border-color: #2563eb !important;
  color: #ffffff !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
}
[data-testid="stDownloadButton"] button:hover {
  background-color: #1d4ed8 !important;
  border-color: #1d4ed8 !important;
}
[data-testid="stDownloadButton"] button p,
[data-testid="stDownloadButton"] button span {
  white-space: nowrap !important;
  color: #ffffff !important;
  margin: 0 !important;
  line-height: 1 !important;
}

/* 모바일: 테이블 숨김, 모바일 전용 영역 표시 */
.dash-mobile-only { display: none; }
@media (max-width: 768px) {
  .dash-wrap { display: none !important; }
  .dash-mobile-only { display: block !important; }
}
/* 데스크톱: 모바일 전용 영역 숨김 */
@media (min-width: 769px) {
  .dash-mobile-only { display: none !important; }
}
</style>
"""


def _build_excel(reqs: list, site_name: str, date_label: str) -> bytes:
    """요청 목록을 엑셀 파일로 변환하여 bytes 반환."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "반입반출현황"

    # ── 스타일 정의 ───────────────────────────────────────────────────────
    thin = Side(style="thin", color="CBD5E1")
    thick = Side(style="medium", color="1E3A8A")
    border_all  = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    hdr_fill   = PatternFill("solid", fgColor="1E3A8A")
    sub_fill   = PatternFill("solid", fgColor="2563EB")
    total_fill = PatternFill("solid", fgColor="E0E7FF")
    even_fill  = PatternFill("solid", fgColor="F8FAFC")

    hdr_font   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    body_font  = Font(name="맑은 고딕", size=10)
    total_font = Font(name="맑은 고딕", bold=True, color="1E3A8A", size=10)
    title_font = Font(name="맑은 고딕", bold=True, size=16)
    meta_font  = Font(name="맑은 고딕", bold=True, size=10, color="1E3A8A")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── 행 1: 제목 ────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"].value = "자재 상·하차 반입/반출 현황"
    ws["A1"].font  = title_font
    ws["A1"].alignment = center
    ws["A1"].border = border_thick
    ws.row_dimensions[1].height = 44

    # ── 행 2: 현장명 / 날짜 ───────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"□ {site_name}"
    ws["A2"].font  = meta_font
    ws["A2"].alignment = left
    ws.merge_cells("G2:L2")
    ws["G2"].value = date_label
    ws["G2"].font  = meta_font
    ws["G2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── 행 3~4: 헤더 (2단) ────────────────────────────────────────────────
    headers_top = [
        ("A", "No"), ("B", "업체명"), ("C", "자재종류"), ("D", "수량"),
        ("E", "반입·반출 차량"), ("F", "상·하차 방식"),
        (None, "장 소"),   # G~H merge
        ("I", "시간"), ("J", "작업지휘자"), ("K", "유도원"), ("L", "담당자"),
    ]
    # 장소 헤더 — G3:H3 merge
    ws.merge_cells("G3:H3")
    ws["G3"].value = "장 소"
    ws["G3"].font  = hdr_font
    ws["G3"].fill  = hdr_fill
    ws["G3"].alignment = center
    ws["G3"].border = border_all

    for col_letter, label in [
        ("A","No"),("B","업체명"),("C","자재종류"),("D","수량"),
        ("E","반입·반출 차량"),("F","상·하차 방식"),
        ("I","시간"),("J","작업지휘자"),("K","유도원"),("L","담당자"),
    ]:
        ws.merge_cells(f"{col_letter}3:{col_letter}4")
        cell = ws[f"{col_letter}3"]
        cell.value = label
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = center
        cell.border = border_all

    # 하위 헤더 — Zone / 장소
    for col_letter, label in [("G","Zone"), ("H","장소")]:
        cell = ws[f"{col_letter}4"]
        cell.value = label
        cell.font  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
        cell.fill  = sub_fill
        cell.alignment = center
        cell.border = border_all

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    # ── 데이터 행 ─────────────────────────────────────────────────────────
    dow_map = {0:"월요일", 1:"화요일", 2:"수요일", 3:"목요일",
               4:"금요일", 5:"토요일", 6:"일요일"}
    total_cnt = 0
    for i, r in enumerate(reqs, 1):
        row_num = i + 4
        kind    = r.get("kind", KIND_IN)
        kind_lbl = "반입" if kind == KIND_IN else "반출"
        company  = r.get("company_name", "")
        item     = r.get("item_name", "")
        vcnt_raw = r.get("vehicle_count", "")
        vcnt     = f"{vcnt_raw}대" if vcnt_raw else ""
        vton     = r.get("vehicle_ton", "")
        loading  = r.get("loading_method", "")
        gate_raw   = r.get("gate", "")
        gate_parts = gate_raw.split("|", 1) if "|" in gate_raw else [gate_raw, ""]
        gate_zone  = gate_parts[0].strip()
        gate_place = gate_parts[1].strip() if len(gate_parts) > 1 else ""
        t_from = r.get("time_from", "")
        sup    = r.get("worker_supervisor", "")
        guide  = r.get("worker_guide", "")
        mgr    = r.get("worker_manager", "")
        vcnt_int = int(vcnt_raw) if str(vcnt_raw).isdigit() else 0
        total_cnt += vcnt_int

        fill = even_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [i, company, item, vcnt, f"{kind_lbl} / {vton}", loading,
                  gate_zone, gate_place, t_from, sup, guide, mgr]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font   = body_font
            cell.fill   = fill
            cell.border = border_all
            cell.alignment = center
        ws.row_dimensions[row_num].height = 18

    # ── 합계 행 ───────────────────────────────────────────────────────────
    total_row = len(reqs) + 5
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"].value = "합 계"
    ws[f"A{total_row}"].font  = total_font
    ws[f"A{total_row}"].fill  = total_fill
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{total_row}"].border = border_all
    ws[f"D{total_row}"].value = f"{total_cnt}대"
    ws[f"D{total_row}"].font  = total_font
    ws[f"D{total_row}"].fill  = total_fill
    ws[f"D{total_row}"].alignment = center
    ws[f"D{total_row}"].border = border_all
    for col_idx in range(5, 13):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill   = total_fill
        cell.border = border_all
    ws.row_dimensions[total_row].height = 18

    # ── 열 너비 ───────────────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter
    col_widths = [5, 14, 14, 7, 14, 12, 10, 14, 8, 12, 10, 10]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── bytes 반환 ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _req_list_for_date(con: sqlite3.Connection, project_id: str, target_date: str):
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute(
        """SELECT * FROM requests
           WHERE project_id=? AND date=?
           ORDER BY time_from, created_at""",
        (project_id, target_date),
    )
    return [dict(r) for r in cur.fetchall()]


def page_dashboard(con: sqlite3.Connection):
    st.markdown(_DASH_CSS, unsafe_allow_html=True)

    project_id = st.session_state.get("PROJECT_ID", "")
    site_name  = settings_get(con, "site_name", "현장명")

    # ── 날짜 상태 ─────────────────────────────────────────────────────────
    if "dash_date" not in st.session_state:
        st.session_state["dash_date"] = date.today()
    cur_date: date = st.session_state["dash_date"]

    # ── 날짜 네비게이션 ───────────────────────────────────────────────────
    with st.container(key="dash_nav"):
        nc1, nc2, nc3, nc4, nc5 = st.columns([1, 1, 3, 1, 1])
        with nc1:
            if st.button("◀◀", key="dash_prev_week", use_container_width=True, help="일주일 전"):
                st.session_state["dash_date"] = cur_date - timedelta(days=7)
                st.rerun()
        with nc2:
            if st.button("◀", key="dash_prev_day", use_container_width=True, help="전날"):
                st.session_state["dash_date"] = cur_date - timedelta(days=1)
                st.rerun()
        with nc3:
            picked = st.date_input(
                "날짜", value=cur_date, key="dash_date_picker",
                label_visibility="collapsed",
            )
            if picked != cur_date:
                st.session_state["dash_date"] = picked
                st.rerun()
        with nc4:
            if st.button("▶", key="dash_next_day", use_container_width=True, help="다음날"):
                st.session_state["dash_date"] = cur_date + timedelta(days=1)
                st.rerun()
        with nc5:
            if st.button("▶▶", key="dash_next_week", use_container_width=True, help="일주일 후"):
                st.session_state["dash_date"] = cur_date + timedelta(days=7)
                st.rerun()

    # ── 데이터 로드 ───────────────────────────────────────────────────────
    target_str = str(cur_date)
    reqs = _req_list_for_date(con, project_id, target_str)

    dow_map = {0:"월요일", 1:"화요일", 2:"수요일", 3:"목요일",
               4:"금요일", 5:"토요일", 6:"일요일"}
    date_label = f"{cur_date.year}년 {cur_date.month}월 {cur_date.day}일 {dow_map[cur_date.weekday()]}"

    # ── 테이블 행 생성 ────────────────────────────────────────────────────
    row_parts = []
    total_cnt = 0

    if not reqs:
        row_parts.append('<tr><td colspan="12" style="padding:30px;color:#94a3b8;">해당 날짜에 등록된 요청이 없습니다.</td></tr>')
    else:
        for i, r in enumerate(reqs, 1):
            kind     = r.get("kind", KIND_IN)
            is_in    = kind == KIND_IN
            kind_cls = "kind-in" if is_in else "kind-out"
            kind_lbl = "반입" if is_in else "반출"
            company  = r.get("company_name", "")
            item     = r.get("item_name", "")
            vcnt_raw = r.get("vehicle_count", "")
            vcnt     = f"{vcnt_raw}대" if vcnt_raw else ""
            vton     = r.get("vehicle_ton", "")
            loading  = r.get("loading_method", "")
            gate_raw   = r.get("gate", "")
            gate_parts = gate_raw.split("|", 1) if "|" in gate_raw else [gate_raw, ""]
            gate_zone  = gate_parts[0].strip()
            gate_place = gate_parts[1].strip() if len(gate_parts) > 1 else ""
            t_from = r.get("time_from", "")
            sup    = r.get("worker_supervisor", "")
            guide  = r.get("worker_guide", "")
            mgr    = r.get("worker_manager", "")
            vcnt_int = int(vcnt_raw) if str(vcnt_raw).isdigit() else 0
            total_cnt += vcnt_int

            tds = (
                f'<td>{i}</td>'
                f'<td>{company}</td>'
                f'<td>{item}</td>'
                f'<td>{vcnt}</td>'
                f'<td><span class="{kind_cls}">{kind_lbl}</span> / {vton}</td>'
                f'<td>{loading}</td>'
                f'<td>{gate_zone}</td>'
                f'<td>{gate_place}</td>'
                f'<td>{t_from}</td>'
                f'<td>{sup}</td>'
                f'<td>{guide}</td>'
                f'<td>{mgr}</td>'
            )
            row_parts.append(f'<tr>{tds}</tr>')

        # 합계 행
        row_parts.append(
            f'<tr class="total-row">'
            f'<td colspan="3" style="text-align:right;padding-right:8px;">합 계</td>'
            f'<td>{total_cnt}대</td>'
            f'<td colspan="8"></td>'
            f'</tr>'
        )

    rows_html = "".join(row_parts)

    # ── HTML 렌더링 ───────────────────────────────────────────────────────
    thead = (
        '<thead>'
        '<tr>'
        '<th rowspan="2">No</th>'
        '<th rowspan="2">업체명</th>'
        '<th rowspan="2">자재종류</th>'
        '<th rowspan="2">수량</th>'
        '<th rowspan="2">반입·반출<br>차량</th>'
        '<th rowspan="2">상·하차<br>방식</th>'
        '<th colspan="2">장 소</th>'
        '<th rowspan="2">시간</th>'
        '<th rowspan="2">작업<br>지휘자</th>'
        '<th rowspan="2">유도원</th>'
        '<th rowspan="2">담당자</th>'
        '</tr>'
        '<tr>'
        '<th class="th-sub">Zone</th>'
        '<th class="th-sub">장소</th>'
        '</tr>'
        '</thead>'
    )
    cnt_summary = f"총 {len(reqs)}건 (반입 {sum(1 for r in reqs if r.get('kind')==KIND_IN)}건 / 반출 {sum(1 for r in reqs if r.get('kind')==KIND_OUT)}건)" if reqs else "등록된 요청 없음"
    mobile_html = (
        f'<div class="dash-mobile-only">'
        f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:20px;text-align:center;">'
        f'<div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:6px;">자재 상·하차 반입/반출 현황</div>'
        f'<div style="font-size:13px;color:#475569;margin-bottom:4px;">□ {site_name}</div>'
        f'<div style="font-size:13px;color:#475569;margin-bottom:12px;">{date_label}</div>'
        f'<div style="font-size:13px;color:#2563eb;font-weight:600;">{cnt_summary}</div>'
        f'</div>'
        f'</div>'
    )
    html = (
        '<div class="dash-wrap">'
        '<div class="dash-title-box"><h2>자재 상 · 하차 반입 / 반출 현황</h2></div>'
        f'<div class="dash-meta"><span>□ {site_name}</span><span>{date_label}</span></div>'
        f'<table class="dash-table">{thead}<tbody>{rows_html}</tbody></table>'
        '</div>'
        f'{mobile_html}'
    )
    st.markdown(html, unsafe_allow_html=True)

    # ── 엑셀 다운로드 버튼 ────────────────────────────────────────────────
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    if reqs:
        try:
            excel_bytes = _build_excel(reqs, site_name, date_label)
            filename = f"반입반출현황_{target_str}.xlsx"
            _, btn_col, _ = st.columns([1.5, 2, 1.5])
            with btn_col:
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        except ImportError:
            st.warning("엑셀 다운로드를 사용하려면 `pip install openpyxl` 을 실행하세요.")
